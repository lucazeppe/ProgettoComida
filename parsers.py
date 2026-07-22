"""Parser per i 3 file di input: Zippi, Amati, Ore lavorate.

Ogni parser ritorna (DataFrame, warnings: list[str]) e solleva ValueError
con messaggio descrittivo per le incoerenze bloccanti.
"""
from __future__ import annotations

import calendar
import re
from datetime import date

import openpyxl
import pandas as pd

import i18n

# Ordine di priorità: "id" (colonna con l'Employee ID 2000xxx aziendale) deve
# vincere su "dipendente"/"empleado" quando entrambe compaiono nella stessa
# intestazione, perché quest'ultime sono spesso un numero dipendente interno
# più corto (es. 100) e NON la chiave di match richiesta.
ID_COLUMN_CANDIDATES_PRIORITY = ["id", "employee id", "id dipendente", "cod ssff"]
ID_COLUMN_CANDIDATES_FALLBACK = ["dipendente", "empleado", "numero de empleado", "employee"]
ID_COLUMN_CANDIDATES = ID_COLUMN_CANDIDATES_PRIORITY + ID_COLUMN_CANDIDATES_FALLBACK


def title_case(s: str) -> str:
    """Ogni parola con sola iniziale maiuscola (gestisce nomi/cognomi con più parole,
    e trattini/apostrofi come separatori, es. "GARCIA-LOPEZ" -> "Garcia-Lopez")."""
    def _cap_word(w: str) -> str:
        parts = re.split(r"([-'])", w)
        return "".join(p if p in ("-", "'") else (p[:1].upper() + p[1:].lower()) for p in parts)
    return " ".join(_cap_word(w) for w in s.split() if w)


def _normalize_id(value) -> str | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, float):
        return str(int(value))
    s = str(value).strip()
    if s == "" or s.lower() == "none":
        return None
    if s.endswith(".0"):
        s = s[:-2]
    return s


def month_business_days(year: int, month: int) -> list[date]:
    """Tutti i giorni Lun-Ven del mese scelto, in ordine."""
    n_days = calendar.monthrange(year, month)[1]
    return [
        d for d in (date(year, month, day) for day in range(1, n_days + 1))
        if d.weekday() < 5
    ]


def month_all_days(year: int, month: int) -> list[date]:
    n_days = calendar.monthrange(year, month)[1]
    return [date(year, month, day) for day in range(1, n_days + 1)]


def _in_month(d: date, year: int, month: int) -> bool:
    return d.year == year and d.month == month


def _parse_hours_value(value) -> float:
    """Converte il valore ore di una cella in float. Gestisce None, numeri,
    stringhe con virgola decimale (es. '8,5') e valori orari (datetime.time)."""
    if value is None:
        return 0.0
    if isinstance(value, str):
        s = value.strip().replace(",", ".")
        if s == "":
            return 0.0
        return float(s)
    if hasattr(value, "hour") and hasattr(value, "minute"):
        return value.hour + value.minute / 60 + getattr(value, "second", 0) / 3600
    return float(value)


# ---------------------------------------------------------------------------
# ZIPPI: righe = dipendente, colonne D..fine = date di ordine (valore datetime)
# ---------------------------------------------------------------------------

def parse_zippi(file, year: int, month: int, warn_weekends: bool = True, lang: str = "it") -> tuple[pd.DataFrame, list[str]]:
    warnings: list[str] = []
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        if row and row[0] and "nombre" in str(row[0]).strip().lower():
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError(i18n.t(lang, "err_zippi_no_header"))

    header = [c for c in ws[header_row_idx]]
    header_values = [c.value for c in header]
    lower_header = [str(v).strip().lower() if v else "" for v in header_values]

    try:
        col_id = lower_header.index("numero de empleado")
    except ValueError:
        try:
            col_id = next(i for i, v in enumerate(lower_header) if "numero de empleado" in v)
        except StopIteration:
            raise ValueError(i18n.t(lang, "err_zippi_no_id_col"))
    try:
        col_email = lower_header.index("correo")
    except ValueError:
        raise ValueError(i18n.t(lang, "err_zippi_no_email_col"))

    date_col_start = max(col_id, col_email) + 1

    records = []
    out_of_month_dates = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        nombre = row[0]
        if nombre is None or str(nombre).strip() == "":
            continue  # riga vuota residua (es. "PRIMERA SEMANA"): salta, non è fine dati
        emp_id = _normalize_id(row[col_id])
        if emp_id is None:
            continue
        for cell in row[date_col_start:]:
            if cell is None:
                continue
            if hasattr(cell, "date"):
                d = cell.date() if hasattr(cell, "date") else cell
            else:
                continue
            if not _in_month(d, year, month):
                out_of_month_dates.append((emp_id, str(nombre), d))
                continue
            records.append({
                "employee_id": emp_id,
                "nombre": str(nombre).strip(),
                "date": d,
            })

    if out_of_month_dates:
        prefix = i18n.t(lang, "on_date_prefix")
        examples = ", ".join(f"{n} {prefix} {d.isoformat()}" for _, n, d in out_of_month_dates[:5])
        raise ValueError(i18n.t(lang, "err_zippi_out_of_month", year=year, month=month, examples=examples))

    df = pd.DataFrame(records, columns=["employee_id", "nombre", "date"])

    if warn_weekends:
        weekend_rows = df[df["date"].apply(lambda d: d.weekday() >= 5)]
        if not weekend_rows.empty:
            warnings.append(i18n.t(lang, "warn_zippi_weekend_orders", n=len(weekend_rows)))

    return df, warnings


# ---------------------------------------------------------------------------
# AMATI: un foglio per settimana, colonne Lunes..Viernes con 0/1
# ---------------------------------------------------------------------------

WEEKDAY_COLS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]


def parse_amati(file, year: int, month: int, lang: str = "it") -> tuple[pd.DataFrame, list[str]]:
    warnings: list[str] = []
    wb = openpyxl.load_workbook(file, data_only=True)

    business_days = month_business_days(year, month)
    weeks: list[list[date]] = []
    current_week: list[date] = []
    for d in business_days:
        if current_week and d.weekday() == 0:
            weeks.append(current_week)
            current_week = []
        current_week.append(d)
    if current_week:
        weeks.append(current_week)
    # Pad ogni settimana a Lun..Ven (5 slot) allineando sul weekday
    padded_weeks = []
    for week in weeks:
        slots: list[date | None] = [None] * 5
        for d in week:
            slots[d.weekday()] = d
        padded_weeks.append(slots)

    # Le settimane richieste si identificano per numero di settimana ISO (non
    # per posizione/ordine delle tab): ogni chunk Lun..Ven del mese ha un solo
    # numero di settimana ISO, uguale per tutti i suoi giorni.
    required_weeks: dict[int, list[date | None]] = {
        week[0].isocalendar()[1]: slots for week, slots in zip(weeks, padded_weeks)
    }

    # Un foglio "appartiene" a una settimana richiesta se quel numero compare
    # tra TUTTI i numeri presenti nel suo nome (non solo il primo): più robusto
    # di un semplice match posizionale. Fogli che non appartengono a nessuna
    # settimana richiesta (numeri extra, es. di un altro mese) sono ignorati.
    sheet_by_week: dict[int, str] = {}
    missing_weeks: list[int] = []
    duplicate_weeks: list[tuple[int, list[str]]] = []
    # Traccia quali settimane richieste "reclama" ciascun foglio: se un foglio
    # contiene più di un numero di settimana richiesto nel nome (es. "Semana
    # 23-24"), matcherebbe correttamente entrambe prese singolarmente (1 solo
    # match ciascuna) senza mai far scattare il controllo duplicate_weeks
    # (che guarda solo "più fogli per la stessa settimana", non il contrario) —
    # duplicando in silenzio gli ordini di quel foglio su due settimane diverse.
    sheet_week_claims: dict[str, list[int]] = {}
    for week_number in sorted(required_weeks):
        matches = [
            name for name in wb.sheetnames
            if week_number in [int(n) for n in re.findall(r"\d+", name)]
        ]
        for name in matches:
            sheet_week_claims.setdefault(name, []).append(week_number)
        if not matches:
            missing_weeks.append(week_number)
        elif len(matches) > 1:
            duplicate_weeks.append((week_number, matches))
        else:
            sheet_by_week[week_number] = matches[0]

    if missing_weeks:
        raise ValueError(i18n.t(
            lang, "err_amati_missing_weeks",
            weeks=", ".join(str(w) for w in missing_weeks), year=year, month=month,
        ))
    if duplicate_weeks:
        details = "; ".join(f"{w} ({', '.join(names)})" for w, names in duplicate_weeks)
        raise ValueError(i18n.t(lang, "err_amati_duplicate_week", details=details))

    ambiguous_sheets = {name: weeks for name, weeks in sheet_week_claims.items() if len(weeks) > 1}
    if ambiguous_sheets:
        name, weeks = next(iter(ambiguous_sheets.items()))
        raise ValueError(i18n.t(
            lang, "err_amati_sheet_multiple_weeks",
            sheet_name=name, weeks=", ".join(str(w) for w in weeks),
        ))

    records = []
    for week_number in sorted(required_weeks):
        sheet_name = sheet_by_week[week_number]
        week_slots = required_weeks[week_number]
        ws = wb[sheet_name]
        header = [c.value for c in ws[1]]
        lower_header = [str(v).strip().lower() if v else "" for v in header]
        try:
            col_nombre = lower_header.index("nombre")
            col_apellidos = lower_header.index("apellidos")
            col_id = lower_header.index("empleado")
        except ValueError as e:
            raise ValueError(i18n.t(lang, "err_amati_bad_header", sheet_name=sheet_name, e=e))

        day_cols = {}
        missing_weekday_cols = []
        for wd_idx, wd_name in enumerate(WEEKDAY_COLS):
            try:
                day_cols[wd_idx] = lower_header.index(wd_name.lower())
            except ValueError:
                missing_weekday_cols.append(wd_name)
        if missing_weekday_cols:
            raise ValueError(i18n.t(
                lang, "err_amati_missing_weekday_cols",
                sheet_name=sheet_name, weekdays=", ".join(missing_weekday_cols),
            ))

        has_data_row = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[col_nombre] is None or str(row[col_nombre]).strip() == "":
                continue
            emp_id = _normalize_id(row[col_id])
            if emp_id is None:
                continue
            has_data_row = True
            nombre = str(row[col_nombre]).strip()
            apellidos = str(row[col_apellidos]).strip() if row[col_apellidos] else ""
            for wd_idx, col_idx in day_cols.items():
                value = row[col_idx]
                if value is None:
                    continue
                if isinstance(value, str):
                    if value.strip() in ("", "0"):
                        continue
                elif not value:
                    continue
                target_date = week_slots[wd_idx]
                if target_date is None:
                    # Giorno del foglio settimanale fuori dal mese scelto (sconfinamento fisiologico)
                    warnings.append(i18n.t(
                        lang, "warn_amati_order_out_of_month",
                        sheet_name=sheet_name, nombre=nombre, apellidos=apellidos,
                        weekday=WEEKDAY_COLS[wd_idx], year=year, month=month,
                    ))
                    continue
                records.append({
                    "employee_id": emp_id,
                    "nombre": nombre,
                    "apellidos": apellidos,
                    "date": target_date,
                })

        if not has_data_row:
            raise ValueError(i18n.t(lang, "err_amati_empty_week_sheet", sheet_name=sheet_name, week=week_number))

    df = pd.DataFrame(records, columns=["employee_id", "nombre", "apellidos", "date"])
    return df, warnings


# ---------------------------------------------------------------------------
# ORE LAVORATE: colonna ID + colonne datetime con ore
# ---------------------------------------------------------------------------

def parse_hours(file, year: int, month: int, warn_weekends: bool = True, lang: str = "it") -> tuple[pd.DataFrame, list[str]]:
    warnings: list[str] = []
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # L'export tipico ha righe di intestazione separate: una riga "titolo" con le
    # date (una per colonna giorno) e un'altra riga con i nomi campo (ID, Nome...).
    # Non si può assumere che siano la stessa riga: si cercano indipendentemente
    # nelle prime righe del file.
    id_row_idx = None
    id_col_idx = None
    for i in range(1, 11):
        row = [c.value for c in ws[i]]
        lower_row = [str(v).strip().lower() if v else "" for v in row]
        # Prima cerca una colonna "ID" prioritaria (nell'ordine di rango della lista,
        # non nell'ordine delle colonne: "id" deve vincere su "cod ssff" anche se
        # quest'ultima compare più a sinistra nell'intestazione); solo se nessuna
        # candidata prioritaria è presente, ripiega su "Dipendente"/"Empleado"
        # (numero interno, non ideale).
        match_j = next((lower_row.index(c) for c in ID_COLUMN_CANDIDATES_PRIORITY if c in lower_row), None)
        if match_j is None:
            match_j = next((lower_row.index(c) for c in ID_COLUMN_CANDIDATES_FALLBACK if c in lower_row), None)
        if match_j is not None:
            id_row_idx = i
            id_col_idx = match_j
            break

    if id_row_idx is None:
        raise ValueError(i18n.t(lang, "err_hours_no_id_col", candidates=ID_COLUMN_CANDIDATES))

    # Nome/Cognome (se presenti) sono nella stessa riga della colonna ID.
    id_header_row = [c.value for c in ws[id_row_idx]]
    lower_id_header_row = [str(v).strip().lower() if v else "" for v in id_header_row]
    nome_col_idx = lower_id_header_row.index("nome") if "nome" in lower_id_header_row else None
    cognome_col_idx = lower_id_header_row.index("cognome") if "cognome" in lower_id_header_row else None

    date_row_idx = None
    best_count = 0
    for i in range(1, 11):
        row = [c.value for c in ws[i]]
        count = sum(1 for v in row if hasattr(v, "date"))
        if count > best_count:
            best_count = count
            date_row_idx = i
    if date_row_idx is None or best_count == 0:
        raise ValueError(i18n.t(lang, "err_hours_no_date_header"))

    date_row = [c.value for c in ws[date_row_idx]]
    all_date_cols = {
        j: v.date() for j, v in enumerate(date_row)
        if hasattr(v, "date")
    }

    header_row_idx = max(id_row_idx, date_row_idx)

    # Il file ore è solo di supporto (serve unicamente a leggere le ore dei
    # giorni del mese scelto): eventuali colonne di altri mesi (es. l'export
    # sconfina nel mese successivo) vengono ignorate silenziosamente, non
    # generano errore né warning.
    date_cols = {j: d for j, d in all_date_cols.items() if _in_month(d, year, month)}
    if not date_cols:
        raise ValueError(i18n.t(lang, "err_hours_no_month_date", year=year, month=month))

    missing_days = sorted(set(month_business_days(year, month)) - set(date_cols.values()))
    if missing_days:
        examples = ", ".join(d.isoformat() for d in missing_days[:5])
        warnings.append(i18n.t(lang, "warn_hours_missing_columns", n=len(missing_days), year=year, month=month, examples=examples))

    records = []
    seen_ids = set()
    duplicate_ids = set()
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if row[id_col_idx] is None:
            continue
        emp_id = _normalize_id(row[id_col_idx])
        if emp_id is None:
            continue
        if not re.fullmatch(r"\d{4,10}", emp_id):
            # Formato ID non riconosciuto: riga ignorata silenziosamente (il file
            # ore è di supporto, non blocca il processamento). L'unico errore
            # bloccante legato all'ID è a valle: un dipendente con ordine ma
            # nessuna corrispondenza nel file ore (vedi business.build_master).
            continue
        if emp_id in seen_ids:
            duplicate_ids.add(emp_id)
        seen_ids.add(emp_id)

        nombre_ore = None
        if nome_col_idx is not None or cognome_col_idx is not None:
            nome_val = row[nome_col_idx] if nome_col_idx is not None and nome_col_idx < len(row) else None
            cognome_val = row[cognome_col_idx] if cognome_col_idx is not None and cognome_col_idx < len(row) else None
            parts = [
                title_case(str(v).strip()) for v in (nome_val, cognome_val)
                if v is not None and str(v).strip() != ""
            ]
            nombre_ore = " ".join(parts) if parts else None

        for j, d in date_cols.items():
            raw_hours = row[j] if j < len(row) else None
            try:
                hours = _parse_hours_value(raw_hours)
            except (ValueError, TypeError):
                raise ValueError(i18n.t(lang, "err_hours_bad_value", raw_hours=raw_hours, emp_id=emp_id, date=d.isoformat()))
            records.append({"employee_id": emp_id, "date": d, "hours": hours, "nombre_ore": nombre_ore})

    if duplicate_ids:
        examples = ", ".join(sorted(duplicate_ids)[:10])
        raise ValueError(i18n.t(lang, "err_hours_duplicate_ids", examples=examples))

    df = pd.DataFrame(records, columns=["employee_id", "date", "hours", "nombre_ore"])

    if warn_weekends:
        weekend_rows = df[(df["hours"] > 0) & (df["date"].apply(lambda d: d.weekday() >= 5))]
        if not weekend_rows.empty:
            warnings.append(i18n.t(lang, "warn_hours_weekend_rows", n=len(weekend_rows)))

    return df, warnings


# ---------------------------------------------------------------------------
# ANAGRAFICA DIPENDENTI: colonna ID + colonna email aziendale, resto ignorato
# ---------------------------------------------------------------------------

def parse_employee_directory(file, lang: str = "it") -> tuple[pd.DataFrame, list[str]]:
    """Export HR con una riga per dipendente: usa solo ID ed email aziendale,
    ignora tutte le altre colonne (reparto, manager, date assunzione, ecc.).
    L'header è sempre alla prima riga (a differenza di Zippi/Ore non serve
    cercarlo su più righe)."""
    warnings: list[str] = []
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header = [c.value for c in ws[1]]
    # Collassa gli spazi multipli prima del confronto: l'export reale ha
    # intestazioni con doppio spazio (es. "Business  Email Information...").
    norm_header = [" ".join(str(v).strip().lower().split()) if v else "" for v in header]

    try:
        col_id = norm_header.index("user/employee id")
    except ValueError:
        raise ValueError(i18n.t(lang, "err_directory_no_id_col"))
    try:
        col_email = norm_header.index("business email information email address")
    except ValueError:
        raise ValueError(i18n.t(lang, "err_directory_no_email_col"))

    records: dict[str, str | None] = {}
    duplicate_ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[col_id] is None:
            continue
        emp_id = _normalize_id(row[col_id])
        if emp_id is None:
            continue
        if not re.fullmatch(r"\d{4,10}", emp_id):
            # Formato ID non riconosciuto: riga ignorata silenziosamente,
            # stesso comportamento del file Ore (file di supporto, non blocca).
            continue
        if emp_id in records:
            duplicate_ids.add(emp_id)
        email = row[col_email]
        records[emp_id] = str(email).strip() if email else None

    if duplicate_ids:
        examples = ", ".join(sorted(duplicate_ids)[:10])
        raise ValueError(i18n.t(lang, "err_directory_duplicate_ids", examples=examples))

    df = pd.DataFrame(
        [{"employee_id": k, "email": v} for k, v in records.items()],
        columns=["employee_id", "email"],
    )
    return df, warnings
