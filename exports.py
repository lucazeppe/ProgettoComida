"""Generazione dei 2 export Excel: solleciti anomalie e riepilogo colorato per fornitore."""
from __future__ import annotations

import calendar
from datetime import date
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from business import ANOMALY_DOUBLE_LABEL_EXPORT, ANOMALY_HOURS_LABEL_EXPORT, is_company_email

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Bianco, Sfondo 1, Più scuro 15%
HEADER_FONT = Font(bold=True)
RED_FONT = Font(color="FF0000")


def _mark_if_non_company_email(ws, row_idx: int, col_idx: int, email) -> None:
    """Colora di rosso la cella email se non ha dominio aziendale."""
    if email and not is_company_email(email):
        ws.cell(row=row_idx, column=col_idx).font = RED_FONT


def build_reminder_export(master_df: pd.DataFrame) -> bytes:
    """ID dipendente, nome, email, giorno anomalia, ragione anomalia.

    Un giorno con entrambe le anomalie genera 2 record distinti.
    master_df deve essere già lo stato effettivo (dopo eventuali forzature).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Anomalies"
    headers = ["Employee ID", "Name", "Email", "Anomaly Date", "Anomaly Reason"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT

    for _, row in master_df.sort_values(["employee_id", "date"]).iterrows():
        email = row["email"] if pd.notna(row["email"]) else ""
        if row["anomaly_double"]:
            ws.append([row["employee_id"], row["nombre"], email, row["date"], ANOMALY_DOUBLE_LABEL_EXPORT])
            _mark_if_non_company_email(ws, ws.max_row, 3, email)
        if row["anomaly_hours"]:
            ws.append([row["employee_id"], row["nombre"], email, row["date"], ANOMALY_HOURS_LABEL_EXPORT])
            _mark_if_non_company_email(ws, ws.max_row, 3, email)

    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = max(12, min(40, max_len + 2))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


MAIL_REMINDER_OBJECT_VALUE = "Comida"

MAIL_REMINDER_HEADERS = (
    ["RecordID", "Employee ID", "Name", "Email", "Object", "Anomaly Hours", "Anomaly Booking",
     "Allegato", "NameAllegato", "ProcessingState"]
    + [f"Att{i}" for i in range(1, 31)]
)


def build_mail_reminder_export(master_df: pd.DataFrame) -> bytes:
    """Un record per dipendente con almeno un'anomalia, per il flusso Power
    Automate di sollecito via mail. Le date di ciascun tipo di anomalia sono
    elencate come lista testuale "GG/MM/AAAA, GG/MM/AAAA" nella stessa cella.
    Il foglio è una vera Tabella Excel (non un semplice range), richiesto da
    Power Automate per leggere correttamente i dati.

    master_df deve essere già lo stato effettivo (dopo eventuali forzature).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Emails"
    ws.append(MAIL_REMINDER_HEADERS)

    anomalies = master_df[master_df["anomaly_hours"] | master_df["anomaly_double"]]
    for record_id, (emp_id, group) in enumerate(
        anomalies.sort_values("date").groupby("employee_id", sort=False), start=1
    ):
        nombre = group["nombre"].iloc[0]
        email = group["email"].iloc[0] if pd.notna(group["email"].iloc[0]) else ""
        hours_dates = group.loc[group["anomaly_hours"], "date"].apply(lambda d: d.strftime("%d/%m/%Y"))
        booking_dates = group.loc[group["anomaly_double"], "date"].apply(lambda d: d.strftime("%d/%m/%Y"))
        ws.append([
            record_id, emp_id, nombre, email, MAIL_REMINDER_OBJECT_VALUE,
            ", ".join(hours_dates), ", ".join(booking_dates),
            "N", None, None,
        ] + [None] * 30)
        _mark_if_non_company_email(ws, ws.max_row, 4, email)

    last_row = max(ws.max_row, 1)
    last_col_letter = get_column_letter(len(MAIL_REMINDER_HEADERS))
    table = Table(displayName="Table1", ref=f"A1:{last_col_letter}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

    for cell in ws[1]:
        cell.font = HEADER_FONT
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 30

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


TOTAL_COLUMN_HEADERS = [
    "Total", "Total Amati", "Total Zippi",
    "# Hours Anomalies", "# Double-Booking Anomalies",
    "Employee Cost", "Note",
]

NOTE_MISSING_HOURS_RECORD = "Hours anomalies due to missing record in Hours file — verify"


def build_summary_export(master_df: pd.DataFrame, year: int, month: int) -> bytes:
    """Righe = dipendenti (ID, nome), colonne = giorni del mese + totali mensili.

    Cella giorno = A / Z / AZ, colorata di verde se eligible quel giorno, rosso
    altrimenti. Dopo le colonne giorno: totali per dipendente (pasti, Amati,
    Zippi, anomalie, costo) e una colonna "Nota" vuota. In fondo, due righe di
    totale giornaliero per fornitore (Amati, Zippi).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month:02d}-{year}"

    n_days = calendar.monthrange(year, month)[1]
    days = [date(year, month, d) for d in range(1, n_days + 1)]
    first_total_col = 3 + len(days)
    col_totale = first_total_col
    col_totale_amati = first_total_col + 1
    col_totale_zippi = first_total_col + 2
    col_costo = first_total_col + TOTAL_COLUMN_HEADERS.index("Employee Cost")
    weekend_col_idxs = {col_idx for col_idx, d in enumerate(days, start=3) if d.weekday() >= 5}

    ws.append(
        ["Employee ID", "Name"] + [d.strftime("%d/%m (%a)") for d in days] + TOTAL_COLUMN_HEADERS
    )
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for col_idx in weekend_col_idxs:
        ws.cell(row=1, column=col_idx).fill = GRAY_FILL

    employees = (
        master_df[["employee_id", "nombre"]]
        .drop_duplicates(subset=["employee_id"])
        .sort_values("nombre")
        .to_dict("records")
    )

    lookup = {
        (row["employee_id"], row["date"]): row
        for _, row in master_df.iterrows()
    }

    for emp in employees:
        emp_id = emp["employee_id"]
        emp_rows = master_df[master_df["employee_id"] == emp_id]
        totale_amati = int(emp_rows["amati"].sum())
        totale_zippi = int(emp_rows["zippi"].sum())
        costo_dipendente = emp_rows["cost_amati"].sum(skipna=True) + emp_rows["cost_zippi"].sum(skipna=True)
        note = NOTE_MISSING_HOURS_RECORD if (emp_rows["anomaly_hours"] & emp_rows["hours_missing"]).any() else None

        line = [emp_id, emp["nombre"]] + [None] * len(days) + [
            totale_amati + totale_zippi,
            totale_amati,
            totale_zippi,
            int(emp_rows["anomaly_hours"].sum()),
            int(emp_rows["anomaly_double"].sum()),
            round(float(costo_dipendente), 2),
            note,
        ]
        ws.append(line)
        excel_row = ws.max_row
        ws.cell(row=excel_row, column=col_costo).number_format = '"$"#,##0.00'
        for col_idx, d in enumerate(days, start=3):
            row = lookup.get((emp_id, d))
            value = None
            if row is not None:
                if row["amati"] and row["zippi"]:
                    value = "AZ"
                elif row["amati"]:
                    value = "A"
                elif row["zippi"]:
                    value = "Z"
            is_weekend = col_idx in weekend_col_idxs
            if value is None and not is_weekend:
                continue
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            if value is not None:
                cell.fill = GREEN_FILL if row["eligible"] else RED_FILL
            elif is_weekend:
                cell.fill = GRAY_FILL
            cell.alignment = Alignment(horizontal="center")

    # Riga vuota di separazione: sotto "Totale" riporta il gran totale pasti
    # mese (Amati+Zippi), una riga sopra le due righe di totale per fornitore.
    last_employee_row = ws.max_row
    ws.append([])
    spacer_row = last_employee_row + 1  # ws.max_row non avanza su un append([]) vuoto
    daily_amati = master_df.groupby("date")["amati"].sum()
    daily_zippi = master_df.groupby("date")["zippi"].sum()
    grand_amati = int(daily_amati.sum())
    grand_zippi = int(daily_zippi.sum())
    ws.cell(row=spacer_row, column=col_totale, value=grand_amati + grand_zippi).font = HEADER_FONT

    ws.append([None, "Total Amati"])
    total_amati_row = ws.max_row
    for col_idx, d in enumerate(days, start=3):
        cell = ws.cell(row=total_amati_row, column=col_idx, value=int(daily_amati.get(d, 0)))
        if col_idx in weekend_col_idxs:
            cell.fill = GRAY_FILL
    # Intersezione riga/colonna Totale Amati = gran totale pasti Amati del mese.
    ws.cell(row=total_amati_row, column=col_totale_amati, value=grand_amati)

    ws.append([None, "Total Zippi"])
    total_zippi_row = ws.max_row
    for col_idx, d in enumerate(days, start=3):
        cell = ws.cell(row=total_zippi_row, column=col_idx, value=int(daily_zippi.get(d, 0)))
        if col_idx in weekend_col_idxs:
            cell.fill = GRAY_FILL
    # Intersezione riga/colonna Totale Zippi = gran totale pasti Zippi del mese.
    ws.cell(row=total_zippi_row, column=col_totale_zippi, value=grand_zippi)

    for r in (total_amati_row, total_zippi_row):
        ws.cell(row=r, column=2).font = HEADER_FONT
    ws.cell(row=total_amati_row, column=col_totale_amati).font = HEADER_FONT
    ws.cell(row=total_zippi_row, column=col_totale_zippi).font = HEADER_FONT

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    for col_idx in range(3, 3 + len(days)):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 10
    for col_idx in range(first_total_col, first_total_col + len(TOTAL_COLUMN_HEADERS)):
        header_len = len(ws.cell(row=1, column=col_idx).value)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, header_len + 2)
    ws.freeze_panes = "C2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
